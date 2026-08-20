import os
from datetime import datetime
from io import BytesIO

import requests
from flask import Flask, Response, jsonify, render_template, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")


def supa_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }


def supa_get(table, params=""):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{table}{params}", headers=supa_headers(), timeout=10)
    r.raise_for_status()
    return r.json()


def supa_insert(table, data):
    r = requests.post(f"{SUPABASE_URL}/rest/v1/{table}", headers=supa_headers(), json=data, timeout=10)
    r.raise_for_status()
    return r.json()


def supa_update(table, data, filters):
    r = requests.patch(
        f"{SUPABASE_URL}/rest/v1/{table}?{filters}",
        headers=supa_headers(),
        json=data,
        timeout=10,
    )
    r.raise_for_status()
    return r.json()


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/asistencia", methods=["GET"])
def get_asistencia():
    records = supa_get("asistencia", "?order=created_at.desc")
    return jsonify(records)


@app.route("/api/marcar", methods=["POST"])
def marcar():
    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    tipo = data.get("tipo")
    if not nombre:
        return jsonify({"error": "El nombre es obligatorio"}), 400
    if tipo not in ("entrada", "salida", "salida_refrigerio", "retorno_refrigerio"):
        return jsonify({"error": "Tipo no válido"}), 400

    now = datetime.now()
    record = {
        "nombre": nombre,
        "tipo": tipo,
        "fecha": now.strftime("%Y-%m-%d"),
        "hora": now.strftime("%H:%M:%S"),
    }
    result = supa_insert("asistencia", record)
    return jsonify({"ok": True, "record": result[0] if result else record})


@app.route("/api/ticket", methods=["POST"])
def crear_ticket():
    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    titulo = (data.get("titulo") or "").strip()
    mensaje = (data.get("mensaje") or "").strip()
    if not nombre or not titulo or not mensaje:
        return jsonify({"error": "Todos los campos son obligatorios"}), 400

    ticket = {
        "nombre": nombre,
        "titulo": titulo,
        "mensaje": mensaje,
        "estado": "abierto",
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    result = supa_insert("tickets", ticket)
    return jsonify({"ok": True, "ticket": result[0] if result else ticket})


@app.route("/api/tickets", methods=["GET"])
def get_tickets():
    tickets = supa_get("tickets", "?order=created_at.desc")
    return jsonify(tickets)


@app.route("/api/ticket/<int:ticket_id>/cerrar", methods=["POST"])
def cerrar_ticket(ticket_id):
    result = supa_update("tickets", {"estado": "cerrado"}, f"id=eq.{ticket_id}")
    if not result:
        return jsonify({"error": "Ticket no encontrado"}), 404
    return jsonify({"ok": True})


@app.route("/api/exportar", methods=["GET"])
def exportar():
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D2D56", end_color="2D2D56", fill_type="solid")
    header_align = Alignment(horizontal="center")
    border = Border(bottom=Side(style="thin", color="CCCCCC"))

    ws1 = wb.active
    ws1.title = "Asistencia"
    ws1.append(["Nombre", "Tipo", "Fecha", "Hora"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    records = supa_get("asistencia", "?order=created_at.desc")
    for r in records:
        ws1.append([r["nombre"], r["tipo"], r["fecha"], r["hora"]])

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.border = border

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 15
    ws1.column_dimensions["D"].width = 12

    ws2 = wb.create_sheet("Tickets")
    ws2.append(["ID", "Nombre", "Asunto", "Mensaje", "Estado", "Fecha"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    tickets = supa_get("tickets", "?order=created_at.desc")
    for t in tickets:
        ws2.append([t["id"], t["nombre"], t["titulo"], t["mensaje"], t["estado"], t["fecha"]])

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.border = border

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=reporte_{fecha}.xlsx"},
    )


if __name__ == "__main__":
    import sys
    port = int(os.environ.get("PORT", 5000))
    debug = "--debug" in sys.argv
    app.run(debug=debug, host="0.0.0.0", port=port)